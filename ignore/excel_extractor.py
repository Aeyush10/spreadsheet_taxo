"""
Comprehensive Excel file extractor that extracts all possible information from Excel spreadsheets.
Supports multiple sheets, macros, charts, images, formulas, formatting, and more.
"""

import os
import json
import pandas as pd
import openpyxl
from openpyxl.drawing.image import Image
import zipfile
import shutil
from pathlib import Path
from typing import Dict, List, Any, Optional
import base64
from datetime import datetime
import warnings
import numpy as np
warnings.filterwarnings('ignore')


class NumpyEncoder(json.JSONEncoder):
    """Custom JSON encoder for numpy data types."""
    def default(self, obj):
        if isinstance(obj, np.integer):
            return int(obj)
        elif isinstance(obj, np.floating):
            return float(obj)
        elif isinstance(obj, np.ndarray):
            return obj.tolist()
        return super(NumpyEncoder, self).default(obj)


class ExcelExtractor:
    """Main class for extracting comprehensive information from Excel files."""
    
    def __init__(self, input_folder: str, output_folder: str):
        """
        Initialize the extractor.
        
        Args:
            input_folder: Path to folder containing Excel files
            output_folder: Path to folder where extracted data will be saved
        """
        self.input_folder = Path(input_folder)
        self.output_folder = Path(output_folder)
        self.output_folder.mkdir(exist_ok=True)
    
    def extract_all_files(self):
        """Extract data from all Excel files in the input folder."""
        excel_files = list(self.input_folder.glob("*.xlsx")) + list(self.input_folder.glob("*.xls"))
        
        if not excel_files:
            print("No Excel files found in the input folder.")
            return
        
        for excel_file in excel_files:
            print(f"Processing: {excel_file.name}")
            try:
                self.extract_file(excel_file)
                print(f"Successfully extracted: {excel_file.name}")
            except Exception as e:
                print(f"Error processing {excel_file.name}: {str(e)}")
    
    def extract_file(self, excel_file: Path):
        """Extract all data from a single Excel file."""
        file_stem = excel_file.stem
        output_dir = self.output_folder / file_stem
        output_dir.mkdir(exist_ok=True)
        
        # Create subdirectories for different types of data
        dirs = {
            'data': output_dir / 'data',
            'charts': output_dir / 'charts',
            'images': output_dir / 'images',
            'metadata': output_dir / 'metadata',
            'formulas': output_dir / 'formulas',
            'styles': output_dir / 'styles',
            'macros': output_dir / 'macros',
            'raw': output_dir / 'raw'
        }
        
        for dir_path in dirs.values():
            dir_path.mkdir(exist_ok=True)
        
        # Load workbook with openpyxl for comprehensive extraction
        wb = openpyxl.load_workbook(excel_file, data_only=False, keep_vba=True)
        
        # Extract basic metadata
        self._extract_metadata(wb, dirs['metadata'])
        
        # Extract data from all sheets
        self._extract_sheet_data(excel_file, dirs['data'])
        
        # Extract formulas
        self._extract_formulas(wb, dirs['formulas'])
        
        # Extract styles and formatting
        self._extract_styles(wb, dirs['styles'])
        
        # Extract images
        self._extract_images(wb, dirs['images'])
        
        # Extract charts
        self._extract_charts(wb, dirs['charts'])
        
        # Extract macros/VBA
        self._extract_macros(wb, dirs['macros'])
        
        # # Extract raw XML structure
        # self._extract_raw_structure(excel_file, dirs['raw'])
        
        # Create summary report
        self._create_summary_report(wb, output_dir)
        
        wb.close()
    
    def _extract_metadata(self, wb: openpyxl.Workbook, output_dir: Path):
        """Extract workbook metadata."""
        metadata = {
            'sheet_names': wb.sheetnames,
            'sheet_count': len(wb.sheetnames),
            'active_sheet': wb.active.title if wb.active else None,
            'properties': {},
            'defined_names': {},
            'creation_date': datetime.now().isoformat(),
            'security': {}
        }
        
        # Extract document properties
        props = wb.properties
        if props:
            metadata['properties'] = {
                'title': props.title,
                'subject': props.subject,
                'creator': props.creator,
                'keywords': props.keywords,
                'description': props.description,
                'last_modified_by': props.lastModifiedBy,
                'created': props.created.isoformat() if props.created else None,
                'modified': props.modified.isoformat() if props.modified else None,
                'category': props.category,
                'content_status': props.contentStatus,
                'version': props.version,
                'revision': props.revision
            }
        
        # Extract defined names
        if wb.defined_names:
            for name in wb.defined_names:
                metadata['defined_names'][name.name] = {
                    'refers_to': str(name.attr_text),
                    'scope': name.localSheetId
                }
        
        # Check for security features
        metadata['security'] = {
            'has_vba': wb.vba_archive is not None,
            'password_protected': False  # This would require additional checking
        }
        
        # Save metadata
        with open(output_dir / 'metadata.json', 'w', encoding='utf-8') as f:
            json.dump(metadata, f, indent=2, ensure_ascii=False, cls=NumpyEncoder)
    
    def _extract_sheet_data(self, excel_file: Path, output_dir: Path):
        """Extract data from all sheets using pandas."""
        # Read all sheets
        all_sheets = pd.read_excel(excel_file, sheet_name=None, engine='openpyxl')
        
        sheet_info = {}
        
        for sheet_name, df in all_sheets.items():
            # Save as CSV
            csv_path = output_dir / f"{sheet_name}.csv"
            df.to_csv(csv_path, index=False, encoding='utf-8')
            
            # Save as JSON
            json_path = output_dir / f"{sheet_name}.json"
            df.to_json(json_path, orient='records', indent=2, force_ascii=False)
            
            # Save as Excel (individual sheet)
            excel_path = output_dir / f"{sheet_name}.xlsx"
            df.to_excel(excel_path, index=False)
            
            # Collect sheet statistics
            sheet_info[sheet_name] = {
                'rows': len(df),
                'columns': len(df.columns),
                'column_names': df.columns.tolist(),
                'data_types': df.dtypes.astype(str).to_dict(),
                'null_counts': df.isnull().sum().to_dict(),
                'memory_usage': df.memory_usage(deep=True).sum()
            }
        
        # Save sheet information
        with open(output_dir / 'sheet_info.json', 'w', encoding='utf-8') as f:
            json.dump(sheet_info, f, indent=2, ensure_ascii=False, cls=NumpyEncoder)
    
    def _extract_formulas(self, wb: openpyxl.Workbook, output_dir: Path):
        """Extract all formulas from the workbook."""
        formulas = {}
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            sheet_formulas = {}
            
            for row in sheet.iter_rows():
                for cell in row:
                    if hasattr(cell, 'column_letter') and cell.data_type == 'f':  # Formula cell
                        cell_ref = f"{cell.column_letter}{cell.row}"
                        sheet_formulas[cell_ref] = {
                            'formula': str(cell.value),
                            'calculated_value': cell.displayed_value,
                            'data_type': cell.data_type
                        }
            
            if sheet_formulas:
                formulas[sheet_name] = sheet_formulas
        
        if formulas:
            with open(output_dir / 'formulas.json', 'w', encoding='utf-8') as f:
                json.dump(formulas, f, indent=2, ensure_ascii=False, cls=NumpyEncoder)
    
    def _extract_styles(self, wb: openpyxl.Workbook, output_dir: Path):
        """Extract styling and formatting information."""
        styles = {}
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            sheet_styles = {}
            
            for row in sheet.iter_rows():
                for cell in row:
                    # Skip merged cells as they don't have column_letter attribute
                    if hasattr(cell, 'column_letter') and cell.has_style:
                        cell_ref = f"{cell.column_letter}{cell.row}"
                        sheet_styles[cell_ref] = {
                            'font': {
                                'name': cell.font.name,
                                'size': cell.font.size,
                                'bold': cell.font.bold,
                                'italic': cell.font.italic,
                                'underline': str(cell.font.underline),
                                'color': str(cell.font.color.rgb) if cell.font.color else None
                            },
                            'fill': {
                                'fill_type': str(cell.fill.fill_type),
                                'start_color': str(cell.fill.start_color.rgb) if cell.fill.start_color else None,
                                'end_color': str(cell.fill.end_color.rgb) if cell.fill.end_color else None
                            },
                            'border': {
                                'top': str(cell.border.top.style) if cell.border.top else None,
                                'right': str(cell.border.right.style) if cell.border.right else None,
                                'bottom': str(cell.border.bottom.style) if cell.border.bottom else None,
                                'left': str(cell.border.left.style) if cell.border.left else None
                            },
                            'alignment': {
                                'horizontal': str(cell.alignment.horizontal),
                                'vertical': str(cell.alignment.vertical),
                                'wrap_text': cell.alignment.wrap_text
                            },
                            'number_format': cell.number_format
                        }
            
            if sheet_styles:
                styles[sheet_name] = sheet_styles
        
        if styles:
            with open(output_dir / 'styles.json', 'w', encoding='utf-8') as f:
                json.dump(styles, f, indent=2, ensure_ascii=False, cls=NumpyEncoder)
    
    def _extract_images(self, wb: openpyxl.Workbook, output_dir: Path):
        """Extract embedded images."""
        images_info = {}
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            sheet_images = []
            
            if hasattr(sheet, '_images'):
                for i, img in enumerate(sheet._images):
                    if isinstance(img, Image):
                        # Save image
                        image_filename = f"{sheet_name}_image_{i+1}.png"
                        image_path = output_dir / image_filename
                        
                        try:
                            # Save the image
                            img.ref.save(str(image_path))
                            
                            sheet_images.append({
                                'filename': image_filename,
                                'anchor': str(img.anchor),
                                'size': f"{img.width}x{img.height}",
                                'format': 'png'
                            })
                        except Exception as e:
                            print(f"Error saving image {image_filename}: {str(e)}")
            
            if sheet_images:
                images_info[sheet_name] = sheet_images
        
        if images_info:
            with open(output_dir / 'images_info.json', 'w', encoding='utf-8') as f:
                json.dump(images_info, f, indent=2, ensure_ascii=False, cls=NumpyEncoder)
    
    def _extract_charts(self, wb: openpyxl.Workbook, output_dir: Path):
        """Extract chart information."""
        charts_info = {}
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            sheet_charts = []
            
            if hasattr(sheet, '_charts'):
                for i, chart in enumerate(sheet._charts):
                    chart_info = {
                        'chart_type': type(chart).__name__,
                        'title': chart.title.tx.rich.p[0].r.t if chart.title and hasattr(chart.title, 'tx') else None,
                        'anchor': str(chart.anchor),
                        'series_count': len(chart.series) if hasattr(chart, 'series') else 0
                    }
                    
                    # Extract series information
                    if hasattr(chart, 'series'):
                        chart_info['series'] = []
                        for series in chart.series:
                            series_info = {
                                'title': str(series.title) if hasattr(series, 'title') else None,
                                'values': str(series.val) if hasattr(series, 'val') else None,
                                'categories': str(series.cat) if hasattr(series, 'cat') else None
                            }
                            chart_info['series'].append(series_info)
                    
                    sheet_charts.append(chart_info)
            
            if sheet_charts:
                charts_info[sheet_name] = sheet_charts
        
        if charts_info:
            with open(output_dir / 'charts_info.json', 'w', encoding='utf-8') as f:
                json.dump(charts_info, f, indent=2, ensure_ascii=False, cls=NumpyEncoder)
    
    def _extract_macros(self, wb: openpyxl.Workbook, output_dir: Path):
        """Extract VBA macros."""
        macros_info = {
            'has_vba': False,
            'vba_modules': []
        }
        
        if wb.vba_archive:
            macros_info['has_vba'] = True
            
            # Save VBA archive
            vba_path = output_dir / 'vba_archive.bin'
            try:
                with open(vba_path, 'wb') as f:
                    # wb.vba_archive is a ZipFile object, we need to get its contents
                    vba_data = wb.vba_archive.read('vbaProject.bin')
                    f.write(vba_data)
                macros_info['vba_archive_saved'] = str(vba_path)
            except Exception as e:
                print(f"Error saving VBA archive: {str(e)}")
                macros_info['vba_archive_error'] = str(e)
        
        with open(output_dir / 'macros_info.json', 'w', encoding='utf-8') as f:
            json.dump(macros_info, f, indent=2, ensure_ascii=False, cls=NumpyEncoder)
    
    def _extract_raw_structure(self, excel_file: Path, output_dir: Path):
        """Extract raw XML structure from Excel file."""
        # Excel files are actually ZIP archives
        with zipfile.ZipFile(excel_file, 'r') as zip_ref:
            # List all files in the archive
            file_list = zip_ref.namelist()
            
            # Save file list
            with open(output_dir / 'file_structure.txt', 'w', encoding='utf-8') as f:
                f.write('\n'.join(file_list))
            
            # Extract key XML files
            key_files = [
                'xl/workbook.xml',
                'xl/app.xml',
                'xl/core.xml',
                'xl/sharedStrings.xml',
                'xl/styles.xml',
                '[Content_Types].xml',
                '_rels/.rels'
            ]
            
            for file_name in key_files:
                if file_name in file_list:
                    try:
                        content = zip_ref.read(file_name)
                        safe_name = file_name.replace('/', '_').replace('[', '').replace(']', '')
                        with open(output_dir / f"{safe_name}", 'wb') as f:
                            f.write(content)
                    except Exception as e:
                        print(f"Error extracting {file_name}: {str(e)}")
    
    def _create_summary_report(self, wb: openpyxl.Workbook, output_dir: Path):
        """Create a comprehensive summary report."""
        summary = {
            'extraction_timestamp': datetime.now().isoformat(),
            'workbook_info': {
                'sheet_count': len(wb.sheetnames),
                'sheet_names': wb.sheetnames,
                'has_vba': wb.vba_archive is not None
            },
            'extracted_components': {
                'data_sheets': len(wb.sheetnames),
                'formulas': 0,
                'images': 0,
                'charts': 0,
                'styles': 0
            },
            'files_created': []
        }
        
        # Count extracted components
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            
            # Count formulas
            for row in sheet.iter_rows():
                for cell in row:
                    if hasattr(cell, 'column_letter') and cell.data_type == 'f':
                        summary['extracted_components']['formulas'] += 1
            
            # Count images
            if hasattr(sheet, '_images'):
                summary['extracted_components']['images'] += len(sheet._images)
            
            # Count charts
            if hasattr(sheet, '_charts'):
                summary['extracted_components']['charts'] += len(sheet._charts)
            
            # Count styled cells
            for row in sheet.iter_rows():
                for cell in row:
                    if hasattr(cell, 'column_letter') and cell.has_style:
                        summary['extracted_components']['styles'] += 1
        
        # List created files
        for item in output_dir.rglob('*'):
            if item.is_file():
                summary['files_created'].append(str(item.relative_to(output_dir)))
        
        with open(output_dir / 'extraction_summary.json', 'w', encoding='utf-8') as f:
            json.dump(summary, f, indent=2, ensure_ascii=False, cls=NumpyEncoder)
        
        print(f"Summary: Extracted {summary['extracted_components']['formulas']} formulas, "
              f"{summary['extracted_components']['images']} images, "
              f"{summary['extracted_components']['charts']} charts from "
              f"{summary['workbook_info']['sheet_count']} sheets")


def main():
    """Main function to run the Excel extractor."""
    input_folder = "raw_spreadsheets"
    output_folder = "spreadsheet_data"
    
    extractor = ExcelExtractor(input_folder, output_folder)
    extractor.extract_all_files()


if __name__ == "__main__":
    main()
