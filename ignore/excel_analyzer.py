"""
Advanced Excel analysis utilities for extracting additional information.
"""

import os
import json
import pandas as pd
import openpyxl
from pathlib import Path
from typing import Dict, List, Any, Optional
import re
from datetime import datetime
import numpy as np


class NumpyEncoder(json.JSONEncoder):
    """Custom JSON encoder for numpy data types."""
    def default(self, obj):
        if isinstance(obj, np.integer):
            return int(obj)
        elif isinstance(obj, np.floating):
            return float(obj)
        elif isinstance(obj, np.ndarray):
            return obj.tolist()
        return super(NumpyEncoder, self).default(obj)


class ExcelAnalyzer:
    """Advanced analyzer for Excel files to extract deeper insights."""
    
    def __init__(self, excel_file: Path):
        """Initialize analyzer with an Excel file."""
        self.excel_file = excel_file
        self.wb = openpyxl.load_workbook(excel_file, data_only=False, keep_vba=True)
        self.wb_data = openpyxl.load_workbook(excel_file, data_only=True)
    
    def analyze_data_patterns(self) -> Dict[str, Any]:
        """Analyze data patterns in all sheets."""
        patterns = {}
        
        for sheet_name in self.wb.sheetnames:
            sheet = self.wb[sheet_name]
            sheet_data = self.wb_data[sheet_name]
            
            # Analyze data types and patterns
            data_analysis = {
                'cell_types': {},
                'data_patterns': {},
                'empty_cells': 0,
                'formula_cells': 0,
                'numeric_cells': 0,
                'text_cells': 0,
                'date_cells': 0,
                'boolean_cells': 0,
                'error_cells': 0
            }
            
            total_cells = 0
            for row in sheet.iter_rows():
                for cell in row:
                    total_cells += 1
                    
                    if cell.value is None:
                        data_analysis['empty_cells'] += 1
                    elif cell.data_type == 'f':
                        data_analysis['formula_cells'] += 1
                    elif cell.data_type == 'n':
                        data_analysis['numeric_cells'] += 1
                    elif cell.data_type == 's':
                        data_analysis['text_cells'] += 1
                    elif cell.data_type == 'd':
                        data_analysis['date_cells'] += 1
                    elif cell.data_type == 'b':
                        data_analysis['boolean_cells'] += 1
                    elif cell.data_type == 'e':
                        data_analysis['error_cells'] += 1
            
            data_analysis['total_cells'] = total_cells
            data_analysis['data_density'] = (total_cells - data_analysis['empty_cells']) / total_cells if total_cells > 0 else 0
            
            patterns[sheet_name] = data_analysis
        
        return patterns
    
    def analyze_formula_dependencies(self) -> Dict[str, Any]:
        """Analyze formula dependencies and references."""
        dependencies = {}
        
        for sheet_name in self.wb.sheetnames:
            sheet = self.wb[sheet_name]
            sheet_deps = {
                'formulas': {},
                'external_references': [],
                'circular_references': [],
                'complex_formulas': []
            }
            
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.data_type == 'f' and cell.value:
                        cell_ref = f"{cell.column_letter}{cell.row}"
                        formula = str(cell.value)
                        
                        # Analyze formula complexity
                        complexity_score = self._calculate_formula_complexity(formula)
                        
                        # Find cell references
                        cell_refs = re.findall(r'[A-Z]+[0-9]+', formula)
                        sheet_refs = re.findall(r'[^!]*![A-Z]+[0-9]+', formula)
                        
                        sheet_deps['formulas'][cell_ref] = {
                            'formula': formula,
                            'complexity_score': complexity_score,
                            'cell_references': cell_refs,
                            'sheet_references': sheet_refs,
                            'functions_used': self._extract_functions(formula)
                        }
                        
                        if complexity_score > 5:
                            sheet_deps['complex_formulas'].append(cell_ref)
                        
                        # Check for external references
                        if '[' in formula and ']' in formula:
                            sheet_deps['external_references'].append(cell_ref)
            
            dependencies[sheet_name] = sheet_deps
        
        return dependencies
    
    def analyze_data_validation(self) -> Dict[str, Any]:
        """Analyze data validation rules."""
        validation = {}
        
        for sheet_name in self.wb.sheetnames:
            sheet = self.wb[sheet_name]
            sheet_validation = {}
            
            if hasattr(sheet, 'data_validations'):
                for dv in sheet.data_validations.dataValidation:
                    for range_obj in dv.cells:
                        range_str = str(range_obj)
                        sheet_validation[range_str] = {
                            'type': dv.type,
                            'formula1': dv.formula1,
                            'formula2': dv.formula2,
                            'showDropDown': dv.showDropDown,
                            'showInputMessage': dv.showInputMessage,
                            'showErrorMessage': dv.showErrorMessage,
                            'errorTitle': dv.errorTitle,
                            'error': dv.error,
                            'promptTitle': dv.promptTitle,
                            'prompt': dv.prompt
                        }
            
            if sheet_validation:
                validation[sheet_name] = sheet_validation
        
        return validation
    
    def analyze_conditional_formatting(self) -> Dict[str, Any]:
        """Analyze conditional formatting rules."""
        conditional_formatting = {}
        
        for sheet_name in self.wb.sheetnames:
            sheet = self.wb[sheet_name]
            sheet_cf = {}
            
            if hasattr(sheet, 'conditional_formatting'):
                for cf in sheet.conditional_formatting:
                    for rule in cf.rules:
                        rule_info = {
                            'type': str(rule.type),
                            'priority': rule.priority,
                            'formula': [str(f) for f in rule.formula] if rule.formula else [],
                            'ranges': [str(r) for r in cf.cells]
                        }
                        
                        if rule.dxf:
                            rule_info['formatting'] = {
                                'font': str(rule.dxf.font) if rule.dxf.font else None,
                                'fill': str(rule.dxf.fill) if rule.dxf.fill else None,
                                'border': str(rule.dxf.border) if rule.dxf.border else None
                            }
                        
                        sheet_cf[f"rule_{rule.priority}"] = rule_info
            
            if sheet_cf:
                conditional_formatting[sheet_name] = sheet_cf
        
        return conditional_formatting
    
    def analyze_pivot_tables(self) -> Dict[str, Any]:
        """Analyze pivot tables."""
        pivot_tables = {}
        
        for sheet_name in self.wb.sheetnames:
            sheet = self.wb[sheet_name]
            sheet_pivots = []
            
            if hasattr(sheet, '_pivots'):
                for pivot in sheet._pivots:
                    pivot_info = {
                        'name': pivot.name if hasattr(pivot, 'name') else None,
                        'cache_id': pivot.cacheId if hasattr(pivot, 'cacheId') else None,
                        'location': str(pivot.location) if hasattr(pivot, 'location') else None
                    }
                    sheet_pivots.append(pivot_info)
            
            if sheet_pivots:
                pivot_tables[sheet_name] = sheet_pivots
        
        return pivot_tables
    
    def analyze_named_ranges(self) -> Dict[str, Any]:
        """Analyze named ranges and their usage."""
        named_ranges = {}
        
        if self.wb.defined_names:
            for name in self.wb.defined_names:
                named_ranges[name.name] = {
                    'refers_to': str(name.attr_text),
                    'scope': name.localSheetId,
                    'comment': getattr(name, 'comment', None),
                    'hidden': getattr(name, 'hidden', False)
                }
        
        return named_ranges
    
    def analyze_protection(self) -> Dict[str, Any]:
        """Analyze workbook and worksheet protection."""
        protection = {
            'workbook_protection': {},
            'sheet_protection': {}
        }
        
        # Workbook protection
        if hasattr(self.wb, 'security'):
            protection['workbook_protection'] = {
                'locked_structure': getattr(self.wb.security, 'lockStructure', False),
                'locked_windows': getattr(self.wb.security, 'lockWindows', False),
                'locked_revision': getattr(self.wb.security, 'lockRevision', False)
            }
        
        # Sheet protection
        for sheet_name in self.wb.sheetnames:
            sheet = self.wb[sheet_name]
            if hasattr(sheet, 'protection'):
                protection['sheet_protection'][sheet_name] = {
                    'sheet_protected': sheet.protection.sheet,
                    'objects_protected': sheet.protection.objects,
                    'scenarios_protected': sheet.protection.scenarios,
                    'format_cells': sheet.protection.formatCells,
                    'format_columns': sheet.protection.formatColumns,
                    'format_rows': sheet.protection.formatRows,
                    'insert_columns': sheet.protection.insertColumns,
                    'insert_rows': sheet.protection.insertRows,
                    'insert_hyperlinks': sheet.protection.insertHyperlinks,
                    'delete_columns': sheet.protection.deleteColumns,
                    'delete_rows': sheet.protection.deleteRows,
                    'select_locked_cells': sheet.protection.selectLockedCells,
                    'sort': sheet.protection.sort,
                    'auto_filter': sheet.protection.autoFilter,
                    'pivot_tables': sheet.protection.pivotTables
                }
        
        return protection
    
    def _calculate_formula_complexity(self, formula: str) -> int:
        """Calculate complexity score for a formula."""
        complexity = 0
        
        # Count nested functions
        complexity += formula.count('(') - formula.count(')')
        
        # Count operators
        operators = ['+', '-', '*', '/', '^', '&', '=', '<', '>', '<=', '>=', '<>']
        for op in operators:
            complexity += formula.count(op)
        
        # Count cell references
        complexity += len(re.findall(r'[A-Z]+[0-9]+', formula))
        
        # Count functions
        functions = re.findall(r'[A-Z]+\(', formula)
        complexity += len(functions)
        
        return complexity
    
    def _extract_functions(self, formula: str) -> List[str]:
        """Extract function names from formula."""
        functions = re.findall(r'([A-Z]+)\(', formula)
        return list(set(functions))
    
    def generate_comprehensive_report(self, output_dir: Path) -> Dict[str, Any]:
        """Generate comprehensive analysis report."""
        report = {
            'analysis_timestamp': datetime.now().isoformat(),
            'file_info': {
                'filename': self.excel_file.name,
                'file_size': self.excel_file.stat().st_size,
                'sheet_count': len(self.wb.sheetnames),
                'sheet_names': self.wb.sheetnames
            },
            'data_patterns': self.analyze_data_patterns(),
            'formula_dependencies': self.analyze_formula_dependencies(),
            'data_validation': self.analyze_data_validation(),
            'conditional_formatting': self.analyze_conditional_formatting(),
            'pivot_tables': self.analyze_pivot_tables(),
            'named_ranges': self.analyze_named_ranges(),
            # 'protection': self.analyze_protection()
        }
        
        # Save comprehensive report
        report_path = output_dir / 'comprehensive_analysis.json'
        with open(report_path, 'w', encoding='utf-8') as f:
            json.dump(report, f, indent=2, ensure_ascii=False, cls=NumpyEncoder)
        
        return report
    
    def close(self):
        """Close the workbook."""
        self.wb.close()
        self.wb_data.close()


def analyze_excel_file(excel_file: Path, output_dir: Path) -> Dict[str, Any]:
    """Analyze a single Excel file and generate comprehensive report."""
    analyzer = ExcelAnalyzer(excel_file)
    try:
        report = analyzer.generate_comprehensive_report(output_dir)
        return report
    finally:
        analyzer.close()


if __name__ == "__main__":
    # Example usage
    input_folder = "raw_spreadsheets"
    output_folder = "spreadsheet_data"

    for excel_file in Path(input_folder).glob("*.xlsx"):
        print(f"Analyzing {excel_file.name}...")
        output_dir = Path(output_folder) / excel_file.stem / "analysis"
        output_dir.mkdir(parents=True, exist_ok=True)
        
        report = analyze_excel_file(excel_file, output_dir)
        print(f"Analysis complete. Report saved to {output_dir / 'comprehensive_analysis.json'}")
