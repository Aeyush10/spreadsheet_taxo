#helper functions - sheet json (formatting removal, hyperlinks, metadata, data validation)
import json
import os
import openpyxl
from sheetjson.core import SheetJson
from sheetjson.util.minisheetjson import minimize_sheetjson_pruned_iterative
from openpyxl.drawing.image import Image
from pathlib import Path
from PIL import Image as PILImage
import io 
import xlwings as xw

INPUT_FOLDER = "orkney_spreadsheets"
OUTPUT_FOLDER = f"{INPUT_FOLDER}_data"

def remove_formatting(sheet_json):
    """
    Remove all formatting from the sheet JSON, keeping only data and structure.
    """
    # Remove 'meta' from overall structure
    if 'meta' in sheet_json:
        del sheet_json['meta']
    
    # Process each worksheet
    for worksheet_name, worksheet_data in sheet_json.get('worksheets', {}).items():
        
        # Remove worksheet properties (contains formatting/display settings)
        if 'worksheetProperties' in worksheet_data:
            del worksheet_data['worksheetProperties']
        
        # Remove formatting from individual cells
        if 'cells' in worksheet_data:
            for cell_ref, cell_data in worksheet_data['cells'].items():
                # Keep only the value, remove all Format information
                if 'Format' in cell_data:
                    del cell_data['Format']
                
                # Remove other formatting-related properties if they exist
                formatting_keys = ['style', 'font', 'fill', 'border', 'alignment', 
                                 'number_format', 'protection']
                for key in formatting_keys:
                    if key in cell_data:
                        del cell_data[key]
        
        #if cells are empty, remove them
        empty_cells = [cell for cell, data in worksheet_data.get('cells', {}).items() if not data or 'value' not in data]
        for cell in empty_cells:
            del worksheet_data['cells'][cell]
        
        # Remove chart formatting but keep chart data structure
        if 'charts' in worksheet_data:
            for chart in worksheet_data['charts']:
                # Remove visual/formatting properties from chart level
                chart_formatting_keys = ['style', 'plotArea', 'chartArea']
                for key in chart_formatting_keys:
                    if key in chart:
                        del chart[key]
                
                # Clean up legend - keep position and visibility, remove formatting
                if 'legend' in chart and isinstance(chart['legend'], dict):
                    legend = chart['legend']
                    # Keep only essential properties
                    essential_legend = {}
                    if 'position' in legend:
                        essential_legend['position'] = legend['position']
                    if 'visible' in legend:
                        essential_legend['visible'] = legend['visible']
                    chart['legend'] = essential_legend
                
                # Clean up title - keep text and essential properties
                if 'title' in chart and isinstance(chart['title'], dict):
                    title = chart['title']
                    essential_title = {}
                    if 'text' in title:
                        essential_title['text'] = title['text']
                    if 'formula' in title:
                        essential_title['formula'] = title['formula']
                    chart['title'] = essential_title
                
                # Remove all formatting from axes
                if 'axes' in chart:
                    for axis_name, axis_data in chart['axes'].items():
                        if isinstance(axis_data, dict):
                            # Keep only essential axis properties, remove all formatting
                            essential_axis = {}
                            
                            # Keep structural/functional properties
                            functional_props = [
                                'position', 'visible', 'numberFormat', 'minimum', 'maximum',
                                'majorUnit', 'minorUnit', 'scaleType', 'categoryType'
                            ]
                            
                            for prop in functional_props:
                                if prop in axis_data:
                                    essential_axis[prop] = axis_data[prop]
                            
                            chart['axes'][axis_name] = essential_axis
                
                # Clean up series - remove formatting but keep data
                if 'series' in chart:
                    for series in chart['series']:
                        # Remove visual formatting from series
                        series_formatting_keys = [
                            'format', 'marker', 'line', 'fill', 'smooth', 'dataLabels',
                            'trendline', 'errorBars', 'pictureOptions'
                        ]
                        for key in series_formatting_keys:
                            if key in series:
                                del series[key]
                        
                        # Keep only essential series data
                        essential_series_props = [
                            'idx', 'order', 'title', 'categories', 'values', 
                            'xValues', 'yValues', 'bubbleSize'
                        ]
                        
                        # Create clean series dict
                        clean_series = {}
                        for prop in essential_series_props:
                            if prop in series:
                                clean_series[prop] = series[prop]
                        
                        # Update the series with clean version
                        series.clear()
                        series.update(clean_series)
        
        # Clean up named items (remove any formatting metadata)
        if 'namedItems' in worksheet_data:
            for named_item in worksheet_data['namedItems']:
                if 'format' in named_item:
                    del named_item['format']
        
        # Remove conditional formatting entirely
        # if 'conditionalFormatting' in worksheet_data:
            # del worksheet_data['conditionalFormatting']
        
        # Remove style-related properties from tables
        if 'tables' in worksheet_data:
            for table in worksheet_data['tables']:
                table_formatting_keys = ['tableStyleInfo', 'format','predefinedTableStyle']
                for key in table_formatting_keys:
                    if key in table:
                        del table[key]
                table_formatting_keys_starters = ['show', 'highlight']
                keys_to_del = []
                for key in table_formatting_keys_starters:
                    for table_key in table:
                        if table_key.startswith(key):
                            keys_to_del.append(table_key)
                for key in keys_to_del:
                    if key in table:
                        del table[key]
    
    return sheet_json

def extract_hyperlinks_from_excel(file_path):
    """
    Extract all hyperlinks from an Excel workbook.
    
    Args:
        file_path: Path to the Excel file
        
    Returns:
        Dictionary with hyperlink information organized by worksheet
    """
    
    workbook = openpyxl.load_workbook(file_path)
    hyperlinks_data = {}
    
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        sheet_hyperlinks = {}

        #get range of rows and columns
        max_row = worksheet.max_row
        max_col = worksheet.max_column

        
        # Check each cell for hyperlinks
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell = worksheet.cell(row=row, column=col)
                if not hasattr(cell.hyperlink, 'target'):
                    continue
                if cell.hyperlink is not None:
                    cell_ref = cell.coordinate
                    hyperlink_info = {
                        'target': cell.hyperlink.target,
                        # 'display': cell.hyperlink.display,
                        # 'tooltip': cell.hyperlink.tooltip,
                        # 'cell_value': cell.value,
                        # 'location': cell.hyperlink.location if hasattr(cell.hyperlink, 'location') else None
                    }
                    sheet_hyperlinks[cell_ref] = hyperlink_info
        
        if sheet_hyperlinks:  # Only add if there are hyperlinks
            hyperlinks_data[sheet_name] = sheet_hyperlinks
        
    
    workbook.close()
    print(hyperlinks_data)
    return hyperlinks_data

def add_hyperlinks_to_sheetjson(sheetjson, file_path):
    """
    Add hyperlink information to the sheetjson structure.
    
    Args:
        sheetjson: The existing sheetjson dictionary
        file_path: Path to the Excel file containing hyperlinks
        
    Returns:
        Updated sheetjson with hyperlinks added
    """
    hyperlinks_data = extract_hyperlinks_from_excel(file_path)
    # Add hyperlinks to each worksheet
    for sheet_name, sheet_data in sheetjson.get('worksheets', {}).items():
        if sheet_name in hyperlinks_data:
            # Add hyperlinks to cells that have them
            sheet_hyperlinks = hyperlinks_data[sheet_name]
            
            for cell_ref, hyperlink_info in sheet_hyperlinks.items():
                # Check if the cell exists in sheetjson
                if 'cells' not in sheet_data:
                    sheet_data['cells'] = {}
                
                if cell_ref not in sheet_data['cells']:
                    sheet_data['cells'][cell_ref] = {}
                
                # Add hyperlink information to the cell
                sheet_data['cells'][cell_ref]['hyperlink'] = hyperlink_info
            
            # Also add a summary of hyperlinks at the worksheet level
            sheet_data['hyperlinks_summary'] = {
                'count': len(sheet_hyperlinks),
                'cells_with_hyperlinks': list(sheet_hyperlinks.keys())
            }
    
    return sheetjson

def add_metadata_to_sheetjson(sheetjson, file_path):
    """
    Add metadata information to the sheetjson structure.
    
    Args:
        sheetjson: The existing sheetjson dictionary
        file_path: Path to the Excel file containing metadata
        
    Returns:
        Updated sheetjson with metadata added
    """
    workbook = openpyxl.load_workbook(file_path)
    
    # Extract metadata from the workbook properties
    #convert properties to a dictionary
    properties = workbook.properties
    properties = {k: str(v) for k, v in properties.__dict__.items() if not k.startswith('_')}
    
    # Add metadata to the json as a new key
    if 'meta' not in sheetjson:
        sheetjson['meta'] = {}
    
    sheetjson['meta'] = properties
    
    return sheetjson

def add_data_validation_to_sheetjson(sheetjson, file_path):
    """
    Extract data validation rules from an Excel file.
    
    Args:
        sheetjson: sheetjson for the excel file
        file_path: Path to the Excel file
    """

    wb = openpyxl.load_workbook(file_path)
    sheet_rules = {}
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        data_validations = sheet.data_validations.dataValidation
        
        if data_validations:
            # Create a dictionary to hold the data validation rules
            validation_rules = {
                'sheet_name': sheet_name,
                'validations': []
            }
            
            for dv in data_validations:
                #rule should remove all None
                rule = {k: str(v) for k, v in dv.__dict__.items() if v is not None}
                validation_rules['validations'].append(rule)
            
            # Save the validation rules to a JSON file
            # output_file = os.path.join(final_output_folder, f"{sheet_name}_data_validation.json")
            # with open(output_file, 'w') as f:
            #     json.dump(validation_rules, f, indent=4)
            
            # print(f"Extracted data validation rules for sheet '{sheet_name}' to {output_file}")
    
            sheet_rules[sheet_name] = validation_rules
    
    # Save all sheet rules to a single JSON file
    for sheet_name in sheet_rules:
        sheetjson['worksheets'][sheet_name]['data_validation'] = sheet_rules[sheet_name]
    
    return sheetjson

# Function to convert an Excel workbook to sheetjson format
def workbook_to_sheetjson(input_folder,input_filename,ouput_folder):
    s2s = SheetJson('openpyxl')
    input_file_path = os.path.join(input_folder,input_filename)
    sheetjson = s2s.fromXLSX(input_file_path)

    #we dont minimise as it deletes some stuff. need to try spreadsheet LLM's compressor
    # sheetjson = minimize_sheetjson_pruned_iterative(sheetjson)

    #dump raw
    # with open('sheetjson.json', 'w') as f:
    #     json.dump(sheetjson, f, indent=4)


    sheetjson_no_formatting = remove_formatting(sheetjson)
    # sheetjson_no_formatting = sheetjson

    sheetjson_no_formatting = add_hyperlinks_to_sheetjson(sheetjson_no_formatting, input_file_path)

    sheetjson_no_formatting = add_metadata_to_sheetjson(sheetjson_no_formatting, input_file_path)

    sheetjson_no_formatting = add_data_validation_to_sheetjson(sheetjson_no_formatting,input_file_path)


    #write the sheetjson to a file
    # with open('sheetjson_no_formatting.json', 'w') as f:
    #     json.dump(sheetjson_no_formatting, f, indent=4)

    output_file_path = os.path.join(ouput_folder, input_filename.split('.')[0])
    os.makedirs(output_file_path, exist_ok=True)
    with open(os.path.join(output_file_path, 'sheetjson.json'), 'w') as f:
        json.dump(sheetjson, f, indent=4)



    #ingore below
    # for sheet_name, sheet_data in sheetjson.get("worksheets", {}).items():
    #         sheet_metadata[sheet_name] = {
    #             'has_tables': len(sheet_data.get('tables', [])) > 0,
    #             'has_charts': len(sheet_data.get('charts', [])) > 0,
    #             'has_pivots': len(sheet_data.get('pivots', [])) > 0,
    #             'has_conditional_formatting': len(sheet_data.get('conditionalFormatting', [])) > 0,
    #             'table_names': [t.get('name') for t in sheet_data.get('tables', [])],
    #             'chart_count': len(sheet_data.get('charts', [])),
    #             'named_items': [ni.get('name') for ni in sheet_data.get('namedItems', [])]
    #     }

#function to extract embedded images from the excel file
def extract_images_from_excel(input_folder, input_file_name, output_folder):
    """
    Extract all embedded images from an Excel workbook and save them to a folder.
    
    Args:
        file_path: Path to the Excel file
        output_folder: Folder to save extracted images
    """
    # Create output folder if it doesn't exist
    Path(output_folder).mkdir(parents=True, exist_ok=True)
    
    # Load the workbook
    input_file_path = os.path.join(input_folder, input_file_name)
    workbook = openpyxl.load_workbook(input_file_path)
    
    image_count = 0
    extracted_images = []
    
    # Iterate through all worksheets
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        
        # Check if the worksheet has any images
        if hasattr(worksheet, '_images') and worksheet._images:
            for img in worksheet._images:
                try:
                    # Get image data
                    image_data = img._data()
                    
                    # Determine file extension based on image format
                    if image_data.startswith(b'\x89PNG'):
                        ext = '.png'
                    elif image_data.startswith(b'\xff\xd8'):
                        ext = '.jpg'
                    elif image_data.startswith(b'GIF'):
                        ext = '.gif'
                    elif image_data.startswith(b'BM'):
                        ext = '.bmp'
                    else:
                        ext = '.png'  # Default to PNG
                    
                    # Create filename
                    filename = f"image{image_count + 1}{ext}"
                    filepath = os.path.join(output_folder, input_file_name.split('.')[0])
                    os.makedirs(filepath, exist_ok=True)
                    filepath = os.path.join(filepath, "images")
                    #CREATE the folder if it does not exist
                    os.makedirs(filepath, exist_ok=True)
                    filepath = os.path.join(filepath, filename)
                    
                    # Save the image
                    with open(filepath, 'wb') as f:
                        f.write(image_data)
                    
                    image_info = {
                        'sheet': sheet_name,
                        'filename': filename,
                        'filepath': filepath,
                        'anchor': getattr(img, 'anchor', None)
                    }
                    
                    extracted_images.append(image_info)
                    image_count += 1
                    
                    print(f"Extracted: {filename} from sheet '{sheet_name}'")
                    
                except Exception as e:
                    print(f"Error extracting image from sheet '{sheet_name}': {e}")
    
    workbook.close()
    
    return extracted_images

#function to extract chart images
def extract_chart_images(input_folder,input_filename, output_folder):
    """
    Opens the specified workbook and returns the first chart object.
    """
    app = xw.App(visible=False)
    input_file_path = os.path.join(input_folder, input_filename)
    wb  = app.books.open(input_file_path)
    chart_output_dir = os.path.join(output_folder, input_filename.split('.')[0])
    os.makedirs(chart_output_dir, exist_ok=True)
    chart_output_dir = os.path.join(chart_output_dir, "charts")
    os.makedirs(chart_output_dir, exist_ok=True)

    for sheet_file in wb.sheets: 
        try:
            
            sht = wb.sheets[sheet_file.name]
            # print(sht.charts)
            i = 1
            for chart in sht.charts:
                # print(chart)
                # print(chart.chart_type)
                chart_filename = f"chart{i}.pdf"
                output_path = os.path.join(chart_output_dir, chart_filename)
                chart.to_pdf(output_path)
                # chart.to_pdf("chart.pdf")
                print(f"Exported chart {i} to {output_path}")
                i += 1
        except Exception as e:  
            print(e) 
            continue 
    
    wb.close()
    app.quit()

from pathlib import Path
#iterate over input folder to get file names
excel_files = list(Path(INPUT_FOLDER).glob("*.xlsx")) + list(Path(INPUT_FOLDER).glob("*.xls"))

for file in excel_files:
    file_name = str(file).split('\\')[-1]
    print("Extracting from file:", file_name)
    # print(file_name)
    workbook_to_sheetjson(INPUT_FOLDER,file_name,OUTPUT_FOLDER)
    extract_images_from_excel(INPUT_FOLDER,file_name,OUTPUT_FOLDER)
    extract_chart_images(INPUT_FOLDER,file_name,OUTPUT_FOLDER)